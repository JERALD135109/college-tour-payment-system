from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from datetime import datetime
import os
import secrets

# -----------------------------
# App Setup
# -----------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(16))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# -----------------------------
# Folder Setup (Render Safe)
# -----------------------------
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
INVOICE_FOLDER = os.path.join(BASE_DIR, "invoices")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(INVOICE_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["INVOICE_FOLDER"] = INVOICE_FOLDER

# -----------------------------
# Excel File
# -----------------------------
EXCEL_FILE = os.path.join(BASE_DIR, "trip_data.xlsx")

ADMIN_PASSWORD = "admin123"

# -----------------------------
# Home Page
# -----------------------------
@app.route("/")
def home():
    return render_template("login.html")

# -----------------------------
# Login (SAFE VERSION)
# -----------------------------
@app.route("/login", methods=["POST"])
def login():
    username = request.form.get("username")
    password = request.form.get("password")

    if not username or not password:
        flash("Invalid login form.")
        return redirect(url_for("home"))

    if password == ADMIN_PASSWORD:
        session.clear()
        session["admin"] = True
        return redirect(url_for("admin_dashboard"))
    else:
        session.clear()
        session["username"] = username
        return redirect(url_for("dashboard"))

# -----------------------------
# User Dashboard
# -----------------------------
@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("home"))

    username = session["username"]

    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        user_data = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                user_data = row
                break

        wb.close()

    except Exception as e:
        return f"Excel Error: {str(e)}"

    return render_template("dashboard.html", user=user_data)

# -----------------------------
# Submit Payment
# -----------------------------
@app.route("/submit_payment", methods=["POST"])
def submit_payment():
    if "username" not in session:
        return redirect(url_for("home"))

    username = session["username"]
    amount = request.form.get("amount")
    file = request.files.get("payment_screenshot")

    if not amount or not file or file.filename == "":
        flash("All fields are required.")
        return redirect(url_for("dashboard"))

    try:
        # Secure filename
        original_filename = secure_filename(file.filename)
        extension = original_filename.split(".")[-1]
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        new_filename = f"{username}_{timestamp}.{extension}"

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], new_filename)
        file.save(filepath)

        # Update Excel
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == username:
                paid_amount = float(row[2].value or 0)
                row[2].value = paid_amount + float(amount)
                row[3].value = "Pending Approval"
                break

        wb.save(EXCEL_FILE)
        wb.close()

        flash("Payment submitted successfully!")

    except Exception as e:
        return f"Payment Error: {str(e)}"

    return redirect(url_for("dashboard"))

# -----------------------------
# Admin Dashboard
# -----------------------------
@app.route("/admin")
def admin_dashboard():
    if "admin" not in session:
        return redirect(url_for("home"))

    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        members = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            members.append(row)

        wb.close()

    except Exception as e:
        return f"Admin Error: {str(e)}"

    return render_template("admin.html", members=members)

# -----------------------------
# Logout
# -----------------------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))

# -----------------------------
# Render PORT Configuration
# -----------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)